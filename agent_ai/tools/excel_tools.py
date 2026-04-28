import logging
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from core.database import SessionLocal
from models.kelompok import Kelompok
from models.mahasiswa import Mahasiswa
from models.kelompokMahasiswa import KelompokMahasiswa
from models.dosen import Dosen
from models.pembimbing import Pembimbing
from models.penguji import Penguji
from models.prodi import Prodi
from models.kategori_pa import KategoriPA
from models.tahun_ajaran import TahunAjaran
from models.tahun_masuk import TahunMasuk

logger = logging.getLogger(__name__)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'storage', 'Kelompok dan Pembimbing.xlsx')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'storage', 'outputs')

# Pastikan output directory ada
os.makedirs(OUTPUT_DIR, exist_ok=True)


def generate_excel_by_context(prodi_id=None, kategori_pa_id=None, tahun_ajaran_id=None, tahun_masuk=None):
    """
    Generate Excel file dengan data kelompok, pembimbing, dan penguji berdasarkan context.
    
    Args:
        prodi_id: Filter by prodi
        kategori_pa_id: Filter by kategori PA
        tahun_ajaran_id: Filter by tahun ajaran
        tahun_masuk: Filter by tahun masuk
        
    Returns:
        dict dengan keys: success (bool), file_path (str), message (str), error (str jika ada)
    """
    session = SessionLocal()
    
    try:
        # Validasi template ada
        if not os.path.exists(TEMPLATE_PATH):
            logger.error(f"Template Excel tidak ditemukan: {TEMPLATE_PATH}")
            return {
                "success": False,
                "error": f"Template Excel tidak ditemukan di {TEMPLATE_PATH}",
                "file_path": None
            }
        
        # Load template
        workbook = load_workbook(TEMPLATE_PATH)
        worksheet = workbook.active

        # Template memiliki merge area untuk layout statis (mis. sampai kelompok 13).
        # Agar export dinamis tetap rapi saat kelompok > 13, normalisasi area data jadi tabel biasa.
        merged_ranges = list(worksheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            if merged_range.max_col < 1 or merged_range.min_col > 7:
                continue
            if merged_range.max_row < 2:
                continue
            worksheet.unmerge_cells(str(merged_range))

        # Bersihkan nilai lama pada area data (A-G, mulai baris 2) agar tidak tercampur dengan isi template.
        for r in range(2, worksheet.max_row + 1):
            for c in range(1, 8):
                worksheet.cell(row=r, column=c).value = None
        
        # ✅ Pastikan header kolom F dan G ada (untuk penguji)
        if worksheet['F1'].value is None:
            worksheet['F1'].value = 'Penguji 1'
        if worksheet['G1'].value is None:
            worksheet['G1'].value = 'Penguji 2'
        
        # Query kelompok sesuai context
        query = session.query(Kelompok)
        
        if prodi_id:
            query = query.filter(Kelompok.prodi_id == prodi_id)
        if kategori_pa_id:
            query = query.filter(Kelompok.KPA_id == kategori_pa_id)
        if tahun_ajaran_id:
            query = query.filter(Kelompok.tahun_ajaran_id == tahun_ajaran_id)
        if tahun_masuk:
            query = query.filter(Kelompok.TM_id == tahun_masuk)
        
        kelompok_list = query.order_by(Kelompok.nomor_kelompok).all()
        
        if not kelompok_list:
            return {
                "success": False,
                "error": "Tidak ada data kelompok sesuai kriteria yang dipilih",
                "file_path": None
            }
        
        # Mulai isi data dari row 2 (row 1 adalah header)
        row = 2
        
        # Set style untuk data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ✅ Kolom F dan G sekarang selalu ada (kita setup di atas)
        has_col_f = True
        has_col_g = True

        def _write_cell(col, row_num, value, align='left'):
            cell = worksheet[f'{col}{row_num}']
            if isinstance(cell, MergedCell):
                return
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal=align)

        for export_group_number, kelompok in enumerate(kelompok_list, start=1):
            # Ambil semua anggota dalam kelompok (bukan hanya 1 orang)
            anggota_list = session.query(Mahasiswa).join(
                KelompokMahasiswa, KelompokMahasiswa.user_id == Mahasiswa.user_id
            ).filter(
                KelompokMahasiswa.kelompok_id == kelompok.id
            ).order_by(Mahasiswa.nim).all()

            # Jika kelompok belum punya anggota, tetap tulis 1 baris agar metadata kelompok terlihat
            if not anggota_list:
                anggota_list = [None]

            # Ambil pembimbing & penguji sekali per kelompok
            pembimbing_1 = session.query(Pembimbing, Dosen).join(
                Dosen, Pembimbing.user_id == Dosen.user_id
            ).filter(
                Pembimbing.kelompok_id == kelompok.id
            ).order_by(Pembimbing.created_at).first()

            pembimbing_1_name = ""
            if pembimbing_1:
                pembimbing_1_name = pembimbing_1.Dosen.nama if pembimbing_1.Dosen else ""

            pembimbing_2 = session.query(Pembimbing, Dosen).join(
                Dosen, Pembimbing.user_id == Dosen.user_id
            ).filter(
                Pembimbing.kelompok_id == kelompok.id
            ).order_by(Pembimbing.created_at).offset(1).first()

            pembimbing_2_name = ""
            if pembimbing_2:
                pembimbing_2_name = pembimbing_2.Dosen.nama if pembimbing_2.Dosen else ""

            if not pembimbing_1_name:
                pembimbing_1_name = "Belum diassign"
            if not pembimbing_2_name:
                pembimbing_2_name = "Belum diassign"

            # ✅ Ambil penguji dengan POLA SAMA seperti pembimbing
            penguji_1 = session.query(Penguji, Dosen).join(
                Dosen, Penguji.user_id == Dosen.user_id
            ).filter(
                Penguji.kelompok_id == kelompok.id
            ).order_by(Penguji.created_at).first()

            penguji_1_name = ""
            if penguji_1:
                penguji_1_name = penguji_1.Dosen.nama if penguji_1.Dosen else ""

            penguji_2 = session.query(Penguji, Dosen).join(
                Dosen, Penguji.user_id == Dosen.user_id
            ).filter(
                Penguji.kelompok_id == kelompok.id
            ).order_by(Penguji.created_at).offset(1).first()

            penguji_2_name = ""
            if penguji_2:
                penguji_2_name = penguji_2.Dosen.nama if penguji_2.Dosen else ""

            if not penguji_1_name:
                penguji_1_name = "Belum diassign"
            if not penguji_2_name:
                penguji_2_name = "Belum diassign"

            # Track row awal kelompok untuk merging
            group_start_row = row
            group_end_row = row + len(anggota_list) - 1

            # Tulis semua anggota. Kolom metadata (C-G) hanya ditulis pada baris pertama kelompok.
            for idx, anggota in enumerate(anggota_list):
                nim = anggota.nim if anggota else ""
                nama = anggota.nama if anggota else ""

                _write_cell('A', row, nim, align='left')
                _write_cell('B', row, nama, align='left')

                # Pembimbing diisi untuk setiap baris anggota agar mudah dibaca di spreadsheet.
                _write_cell('D', row, pembimbing_1_name, align='left')
                _write_cell('E', row, pembimbing_2_name, align='left')

                # ✅ Penguji juga diisi untuk setiap baris agar mudah dibaca
                _write_cell('F', row, penguji_1_name, align='left')
                _write_cell('G', row, penguji_2_name, align='left')

                if idx == 0:
                    # Nomor kelompok pada export dibuat urut agar bila >13 otomatis lanjut 14, 15, dst.
                    _write_cell('C', row, export_group_number, align='center')

                row += 1

            # ✅ MERGE CELLS untuk kelompok dengan multiple anggota
            if group_start_row < group_end_row:
                # Merge kolom C (Kelompok)
                worksheet.merge_cells(f'C{group_start_row}:C{group_end_row}')
                worksheet[f'C{group_start_row}'].alignment = Alignment(horizontal='center', vertical='center')

                # Merge kolom D (Pembimbing 1)
                worksheet.merge_cells(f'D{group_start_row}:D{group_end_row}')
                worksheet[f'D{group_start_row}'].alignment = Alignment(horizontal='left', vertical='center')

                # Merge kolom E (Pembimbing 2)
                worksheet.merge_cells(f'E{group_start_row}:E{group_end_row}')
                worksheet[f'E{group_start_row}'].alignment = Alignment(horizontal='left', vertical='center')

                # ✅ Merge kolom F (Penguji 1) - selalu ada sekarang
                worksheet.merge_cells(f'F{group_start_row}:F{group_end_row}')
                worksheet[f'F{group_start_row}'].alignment = Alignment(horizontal='left', vertical='center')

                # ✅ Merge kolom G (Penguji 2) - selalu ada sekarang
                worksheet.merge_cells(f'G{group_start_row}:G{group_end_row}')
                worksheet[f'G{group_start_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Sesuaikan lebar kolom berdasarkan konten agar lebih mudah dibaca.
        for col in worksheet.iter_cols(min_col=1, max_col=worksheet.max_column):
            col_letter = col[0].column_letter
            max_length = 0
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            # Batasi lebar agar tetap rapi pada kolom yang berisi teks panjang.
            worksheet.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 60)

        # Generate filename dengan timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Kelompok_Pembimbing_Penguji_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename)
        
        # Save file
        workbook.save(output_path)
        logger.info(f"Excel file generated: {output_path}")
        
        return {
            "success": True,
            "file_path": output_path,
            "filename": filename,
            "message": f"Excel berhasil dibuat dengan {len(kelompok_list)} kelompok",
            "row_count": len(kelompok_list)
        }
        
    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")
        return {
            "success": False,
            "error": f"Error: {str(e)}",
            "file_path": None
        }
    finally:
        session.close()


def get_excel_output_filename(file_path):
    """Extract filename dari full path"""
    if file_path:
        return os.path.basename(file_path)
    return None
